import git
import os
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

# Function to check if a commit is within the date range
def is_commit_in_range(commit_date, from_date, to_date):
    # Convert commit_date to naive (remove timezone info if it's aware)
    if commit_date.tzinfo is not None:
        commit_date = commit_date.replace(tzinfo=None)  # Convert to naive datetime

    # Convert from_date and to_date to naive as well
    if from_date.tzinfo is not None:
        from_date = from_date.replace(tzinfo=None)  # Convert to naive datetime
    if to_date.tzinfo is not None:
        to_date = to_date.replace(tzinfo=None)  # Convert to naive datetime

    return from_date <= commit_date <= to_date

# Function to update Excel sheet
def update_excel(sheet, row, person_name, repo_url, commit_in_range, date_ranges):
    sheet[f"A{row}"] = person_name
    sheet[f"B{row}"] = repo_url
    
    # For each date range, check if any commit falls within the range and update accordingly
    for i, date_range in enumerate(date_ranges, start=1):
        from_date, to_date = date_range
        commits_in_range = commit_in_range.get(i, [])
        
        # If there are commits in the range, write each commit time in the column
        if commits_in_range:
            commit_times = ", ".join(commits_in_range)
            sheet.cell(row=row, column=i+2, value=f"Yes ({commit_times})")
        else:
            sheet.cell(row=row, column=i+2, value="No")

# Function to process a repository and check against multiple date ranges
def process_repo(person_name, repo_url, date_ranges, sheet, row):
    try:
        # Clone the repo
        repo_name = repo_url.split("/")[-1].replace(".git", "")
        clone_dir = f"./{repo_name}"
        
        if not os.path.exists(clone_dir):
            print(f"Cloning {repo_url}...")
            git.Repo.clone_from(repo_url, clone_dir)
        
        # Get commit history and check the date range
        repo = git.Repo(clone_dir)
        commit_in_range = {}  # Dictionary to store commits that fall in each range
        
        for i, (from_date, to_date) in enumerate(date_ranges, start=1):
            commit_in_range[i] = []  # Initialize an empty list for each range
            for commit in repo.iter_commits():
                commit_date = commit.committed_datetime
                if is_commit_in_range(commit_date, from_date, to_date):
                    commit_in_range[i].append(commit_date.strftime("%H:%M:%S"))  # Add commit time to the list
        
        # Update the Excel sheet with the results for the current repo
        update_excel(sheet, row, person_name, repo_url, commit_in_range, date_ranges)
    except Exception as e:
        print(f"Error processing {repo_url}: {e}")

# Main function to process and update Excel
def main():
    # Define the multiple date ranges for commit check, including the new date range
    date_ranges = [
        (datetime(2023, 1, 1), datetime(2023, 4, 1)),  # Range 1
        (datetime(2023, 3, 31), datetime(2023, 6, 30)),  # Range 2
        (datetime(2024, 11, 20), datetime(2024, 11, 25))
        # Additional ranges can be added here
    ]

    # Load or create Excel sheet
    excel_file = "repo_info.xlsx"
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "GitHub Repos"
        sheet.append(["Person Name", "GitHub Repo URL", "1/1/2023 to 4/1/2023", "3/31/2023 to 6/30/2023"])
    else:
        wb = load_workbook(excel_file)
        sheet = wb.active
    
    # List of GitHub repos to process (with person name and repo URL)
    repo_info = [
        {"person_name": "abhi", "repo_url": "https://github.com/Abhi2971/APJFSA_N.git"},
        {"person_name": "pooja", "repo_url": "https://github.com/pjyadav321/APJFSA_N.git"},
        # Add more repositories here
    ]
    
    # Process each repository
    for i, info in enumerate(repo_info):
        process_repo(info["person_name"], info["repo_url"], date_ranges, sheet, i+2)
    
    # Save the updated Excel file
    wb.save(excel_file)
    print("Excel file updated.")

if __name__ == "__main__":
    main()
# my new code of a1

